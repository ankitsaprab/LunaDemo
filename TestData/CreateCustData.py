import openpyxl
import pytest


class userdata:

    Detail = [{"first":"Tom", "last": "Cick", "mail":"mail@mail.com", "pass":"Sipl@112"}, {"first":"Amar", "last": "akbar", "mail":"mail@mail.com", "pass":"Sipl@112"}]

    @staticmethod
    def exldata():
        book = openpyxl.load_workbook("D:\\PycharmProjects\\Luma\\TestdataE.xlsx")
        sheet = book.active
        Dic = {}
        for i in range(2, sheet.max_row + 1):
            for j in range(2, sheet.max_column + 1):
                # if sheet.cell(row=i, column=1).value == "test_CheckWithDuplicatEmail_02":
                Dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        
        return [Dic]
            # print(Dic)


    # ("Amar", "Akbar", "Anthony")

    # @pytest.fixture(params=userdata.Detail)
    # def getUserData(self, request):
    #     return request.params
